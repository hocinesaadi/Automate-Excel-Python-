from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
wb = load_workbook('NewGrades.xlsx')

ws = wb.active
#change Value
ws['A2'].value = "Test"

wb.save('NewGrades.xlsx')


#create new sheet
wb.create_sheet("Test")
print(wb.sheetnames)

#Accessing Multiple Cells
for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)


#Merging Cells 
ws.merge_cells('A1:D1')

ws.unmerge_cells('A1:D1')

#Inserting And Deleting Rows,Columns
ws.insert_rows(7)
ws.insert_cols(2)
ws.delete_rows(7)
ws.delete_cols(2)

#Copying And Moving Cells
ws.move_range("C1:D11",rows=2,cols=2)



